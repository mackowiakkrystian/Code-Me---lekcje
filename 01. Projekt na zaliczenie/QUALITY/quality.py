from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import progressbar
import os

# BIBLIOTEKA DO SORTOWANIA

# LICZYMY CZAS W SEKUNDACH WYKONYWANIA KODU PART 1
import time

start_time = time.time()

plik_wsadowy = input("\n\nPodaj nazwe pliku z danymi wsadowymi: (format XLSX) ")

# SORTOWANIE ON
fileDir = os.path.dirname(os.path.realpath('__file__'));
filename = os.path.join(fileDir, './' + plik_wsadowy + '.xlsx')
print("\nZaladowano plik: " + filename + "\n")



# to komendy od OPENPYXL
book = load_workbook(str(plik_wsadowy) +'.xlsx')
zakladkaall = book.get_sheet_by_name('Sheet1') # koniec komend





close = "no"
while close == "no":

    licznik = 1
    liczkolumny = 1

    tablicaall = {
        licznik: {liczkolumny: ""}
    }

    for wiersz in range(2, zakladkaall.max_row + 1):
        for kolumna in "ABCDEFGHIJKLMNOPQRST":
            komorka = "{}{}".format(kolumna, wiersz)
            wartosckomorki = zakladkaall[komorka].value

            if liczkolumny == 1:
                tablicaall[licznik] = {}

            tablicaall[licznik][liczkolumny] = wartosckomorki

            if liczkolumny < 20:
                liczkolumny += 1
            else:
                liczkolumny = 1
                licznik += 1

    #### DODAJ ZAKLADKE ASSETS I WYPISZ DANE ####
    assets = {}
    assets2 = {}
    counter = 1

    for keyasset in tablicaall:
        ass = tablicaall[keyasset]
        for jj in ass:
            if jj == 1:
                assets[keyasset] = {}

            assets[keyasset][jj] = ass[jj]

        for hh in ass:
            if hh == 1:
                assets2[counter] = {}
            assets2[counter][hh] = ass[hh]
        counter += 1

    book.create_sheet("Analysis", 2)
    result_tab = book.get_sheet_by_name('Analysis')

    style_tresc = Font(bold=False, italic=False, name='Helvetica-Normal', size=7, color='737373')
    style_blad_bold = Font(bold=True, italic=False, name='Helvetica-Normal', size=7, color='A80000')
    style_blad = Font(bold=False, italic=False, name='Helvetica-Normal', size=7, color='A80000')

    kolor = 0
    next_row2 = 1
    numerkolumny = 1
    assets3 = {}
    print("Sprawdzanie: 4 warunków")

    for zmienna2 in progressbar.progressbar(assets2, redirect_stdout=True):
        assets3 = assets2[zmienna2]
        for w2, s2 in assets3.items():
            if next_row2 == 1:
                result_tab.cell(column=1, row=1, value="Reference")
                result_tab.cell(column=2, row=1, value="Doc number")
                result_tab.cell(column=3, row=1, value="Doc type")
                result_tab.cell(column=4, row=1, value="Block")
                result_tab.cell(column=5, row=1, value="Vendor account")
                result_tab.cell(column=6, row=1, value="Text")
                result_tab.cell(column=7, row=1, value="User name")
                result_tab.cell(column=8, row=1, value="Entry date")
                result_tab.cell(column=9, row=1, value="Doc date")
                result_tab.cell(column=10, row=1, value="Amount in doc. curr.")
                result_tab.cell(column=11, row=1, value="Doc Currency")
                result_tab.cell(column=12, row=1, value="Amount in local curr.")
                result_tab.cell(column=13, row=1, value="Local currency")
                result_tab.cell(column=14, row=1, value="Net due date")
                result_tab.cell(column=15, row=1, value="Comments")
                result_tab.cell(column=16, row=1, value="Clearing?")
                result_tab.cell(column=17, row=1, value="Currency check")

                next_row2 += 1

            # 1. TUTAJ SPRAWDZAMY CZY DATA ZAPOSTOWANIA JEST WCZESNIEJSZA
            # NIZ DATA FAKTURY. JESLI TAK - ERROR
            if (assets3[13] < assets3[14]):
                result_tab.cell(column=15, row=next_row2,
                                value="Data dokumentu jest pozniejsza niz zapostowania!").font = style_blad_bold
                kolor = 1

            # 2. TUTAJ SPRAWDZAMY GOODYEAR'A - KREDYT NOTA CZY ZOSTALA ZAPOSTOWANA W TYM SAMYM MIESIACU
            if (str(assets3[6]) == "107372"):  # jesli to goodyear
                if (assets3[15] > 0):  # jesli to credit note

                    miesiac_z_blinedate = assets3[14]  # wyciagamy miesiac z doc date
                    dzisiejszy_miesiac = datetime.now()  # wyciagamy miesiac z daty dzisiejszej

                    if (dzisiejszy_miesiac.month == miesiac_z_blinedate.month):  # jesli to kredyt nota z miesiaca platnosci - blok
                        if (assets3[5] is None):
                            # wypisz error jesli warunki dla goodyear'a sa spelnione
                            result_tab.cell(column=15, row=next_row2,
                                            value="GoodYear - zablokowac kredyt note!").font = style_blad_bold
                            kolor = 1

            # 3. TUTAJ SPRAWDZAMY CZY TO AIRPLUS
            if (str(assets3[6]) == "108529"):  # jesli to airplus
                if (assets3[5] is None):  # jesli nie ma blokady
                    # wypisz error jesli warunki dla airplus sa spelnione
                    result_tab.cell(column=15, row=next_row2, value="Airplus - zablokowac!").font = style_blad_bold
                    kolor = 1

            # 4. TUTAJ SPRAWDZAMY CZY TO ALLSTAR
            if (str(assets3[6]) == "100712"):  # jesli to allstar
                if (assets3[5] is None):  # jesli nie ma blokady
                    # wypisz error jesli warunki dla allstar sa spelnione
                    result_tab.cell(column=15, row=next_row2, value="Allstar - zablokowac!").font = style_blad_bold
                    kolor = 1

            if (w2 == 2) or (w2 == 3) or (w2 == 4) or (w2 == 5) or (w2 == 6) or (w2 == 10) or (w2 == 12) or (
                w2 == 13) or (w2 == 14) or (w2 == 15) or (w2 == 16) or (w2 == 17) or (w2 == 18) or (w2 == 20):
                # wypisz odpowiednie kolumny (nie wszystkie, dlatego tyle warunkow powyzej!
                # zmien kolor na szary jesli nie ma erroru

                if kolor == 1:  # error - zmien kolor czerwony
                    if (w2 == 15) or (w2 == 17):  # jesli natrafisz na kwoty - konwert na 00.00
                        result_tab.cell(column=numerkolumny, row=next_row2, value=assets3[w2]).font = style_blad
                        result_tab.cell(column=numerkolumny, row=next_row2).number_format = '0.00'
                        result_tab.row_dimensions[next_row2].height = 11
                        numerkolumny = numerkolumny + 1

                    elif (w2 == 13) or (w2 == 14) or (
                        w2 == 20):  # jesli to kolumna z data, skonwertuj aby nie bylo godzin
                        datka = assets3[w2]
                        result_tab.cell(column=numerkolumny, row=next_row2,
                                        value=datka.strftime("%d.%m.%Y")).font = style_blad
                        result_tab.row_dimensions[next_row2].height = 11
                        numerkolumny = numerkolumny + 1
                    else:
                        result_tab.cell(column=numerkolumny, row=next_row2, value=assets3[w2]).font = style_blad
                        result_tab.row_dimensions[next_row2].height = 11
                        numerkolumny = numerkolumny + 1

                else:  # brak erroru - kolor szary
                    if (w2 == 15) or (w2 == 17):  # jesli natrafisz na kwoty - konwert na 00.00
                        result_tab.cell(column=numerkolumny, row=next_row2, value=assets3[w2]).font = style_tresc
                        result_tab.cell(column=numerkolumny, row=next_row2).number_format = '0.00'
                        result_tab.row_dimensions[next_row2].height = 11
                        numerkolumny = numerkolumny + 1

                    elif (w2 == 13) or (w2 == 14) or (
                        w2 == 20):  # jesli to kolumna z data, skonwertuj aby nie bylo godzin
                        datka = assets3[w2]
                        result_tab.cell(column=numerkolumny, row=next_row2,
                                        value=datka.strftime("%d.%m.%Y")).font = style_tresc
                        result_tab.row_dimensions[next_row2].height = 11
                        numerkolumny = numerkolumny + 1
                    else:
                        result_tab.cell(column=numerkolumny, row=next_row2, value=assets3[w2]).font = style_tresc
                        result_tab.row_dimensions[next_row2].height = 11
                        numerkolumny = numerkolumny + 1

                kolor = 0  # zresetuj kod bledu

            if w2 == 20:
                next_row2 += 1
                numerkolumny = 1

    # WYPISANIE ELEMENTOW GOTOWYCH DO CLEARINGU
    next_row2 = 1
    numerkolumny = 1
    nastepny = 1
    numer_pary = 1

    print("Sprawdzanie: gotowe elementy do clearingu")

    for zmienna2 in assets2:

        # JESLI MA W OPISIE 'MR8M' LUB 'COCKPIT'
        if (assets2[zmienna2][10] == "MR8M") or (assets2[zmienna2][10] == "/COCKPIT/1"):
            nastepny = 1

            while nastepny <= len(assets2):
                kwota = assets2[nastepny][15] * (-1)  # KONWERT KWOTY ODWROTNEJ NA TA SAMA - MINUS 1

                if (assets2[zmienna2][2] == assets2[nastepny][2]) and (nastepny != zmienna2) and (
                    assets2[zmienna2][6] == assets2[nastepny][6]) and (
                    assets2[zmienna2][14] == assets2[nastepny][14]) and ((assets2[zmienna2][15] == kwota) or (
                            assets2[zmienna2][15] == assets[nastepny][15])):
                    # POWYZEJ: SPRAWDZENIE WARUNKOW (CZY TA SAMA REFERENCJA, DATA, KWOTA, KONTO)

                    result_tab.cell(column=16, row=zmienna2+1, value="Do CLEARINGU. PARA NR.: "+str(numer_pary)).font = style_blad_bold
                    result_tab.cell(column=16, row=nastepny+1, value="Do CLEARINGU. PARA NR.: "+str(numer_pary)).font = style_blad_bold
                    nastepny = nastepny + 1

                else:
                    nastepny = nastepny + 1

                if nastepny == len(assets2):  # licznik pary + 1, jak znajdzie wszystkie itemy pasujace z Excela
                    numer_pary = numer_pary + 1



    # SPRAWDZANIE WALUTY DLA DANEGO VENDORA - CZY SA ODCHYLY?

    ilosc_gbp = 0
    ilosc_inne = 0
    dlugosctablicy = len(assets2)
    iteracja = 1
    wiersz = 0
    start = 1

    print("Sprawdzanie: czy sa odchyly w walucie")

    while wiersz < dlugosctablicy:
        wiersz += 1
        for komorka in range(1, dlugosctablicy+1):
            if iteracja < dlugosctablicy:
                if assets2[wiersz][6] == assets2[iteracja][6]:
                    if assets2[iteracja][16] == "GBP":
                        ilosc_gbp = ilosc_gbp + 1

                    else:
                        ilosc_inne = ilosc_inne + 1
                    iteracja = iteracja + 1

                else: # JESLI TO NIE JEST OSTATNI ELEMENT - WYKONUJ PONIZSZY KOD
                    stop = iteracja-1
                    suma_walut = ilosc_gbp + ilosc_inne

                    if ((ilosc_gbp != suma_walut) and (ilosc_inne != suma_walut)):
                        for wiersz2 in range(start, stop+1):
                            result_tab.cell(column=17, row=wiersz2+1, value="Check Currency in vendor " +assets2[wiersz2][6]).font = style_blad_bold

                    start = stop+1
                    komorka = stop
                    wiersz = iteracja
                    wiersz2 = iteracja
                    ilosc_gbp = 0
                    ilosc_inne = 0

            else:     # TO JEST OSTATNI ELEMENT TABLICY
                if wiersz <= dlugosctablicy:
                    if assets2[wiersz][6] == assets2[dlugosctablicy][6]:
                        if assets2[dlugosctablicy][16] == "GBP":
                            ilosc_gbp = ilosc_gbp + 1

                        else:
                            ilosc_inne = ilosc_inne + 1

                        wiersz += 1
                    else:
                        stop = dlugosctablicy-1
                        suma_walut = ilosc_gbp + ilosc_inne

                        if ((ilosc_gbp != suma_walut) and (ilosc_inne != suma_walut)):
                            for wiersz2 in range(start+1, stop+1):
                                result_tab.cell(column=17, row=start+2, value="Check Currency in vendor " +assets2[wiersz2][6]).font = style_blad_bold
                                start += 1


                    if wiersz == dlugosctablicy: # JESLI TO JEST OSTATNI ELEMENT - WYPISZ I KONIEC
                        stop = dlugosctablicy
                        suma_walut = ilosc_gbp + ilosc_inne

                        if ((ilosc_gbp != suma_walut) and (ilosc_inne != suma_walut)):
                            for wiersz2 in range(start, stop+1):
                                result_tab.cell(column=17, row=start+1, value="Check Currency in vendor " +assets2[wiersz2][6]).font = style_blad_bold
                                start += 1


    # DEFINICJA STYLI DLA ZAKLADKI DANE:
    naglowek_data_tab = Font(bold=True, italic=False, name='Arial', size=14, color='000000')
    style_data_tab = Font(bold=False, italic=False, name='Arial', size=14, color='000000')
    numerydok_data_tab = Font(bold=True, italic=False, name='Arial', size=14, color='FFFFFF')
    numerydok_data_tab_kolor = PatternFill(start_color='003466', end_color='003466', fill_type='solid')
    naglowek_data_tab_kolor = PatternFill(start_color='d4d6d9', end_color='d4d6d9', fill_type='solid')

    # STWORZ KOLEJNA ZAKLADKE - LATWE KOPIOWANIE DANYCH DO COM5 SPM
    book.create_sheet("DATA READY TO COPY", 2)
    data_tab = book.get_sheet_by_name('DATA READY TO COPY')
    data_tab.sheet_view.zoomScale = 40
    print("Utworzono zakladke 'Data ready to copy'")

    # OTWIERAMY PLIK TXT Z ZAPISANYMI NUMERAMI FAKTUR
    pliktxt = {}
    nr = 1
    plikjestpusty = 1

    try:
        k = open("dokumenty.txt")
        plikjestpusty = 0
        k.close()
    except IOError:
        plikjestpusty = 1

    if plikjestpusty == 0:
        f = open("dokumenty.txt", 'a+')
    else:
        f = open("dokumenty.txt", 'w+')

    with open("dokumenty.txt") as f:
        for line in f:
            (val) = line.split()
            pliktxt[nr] = val
            nr = nr + 1


    dict_comments = {}
    analysis_tab = book.get_sheet_by_name('Analysis')
    iter = 1
    iter_column = 1
    nr_wiersza = 2
    wolniej = 1
    szybciej = 1
    blad_data = 0
    blad_airplus = 0
    blad_allstar = 0
    blad_currency = 0

    zladata = "Data dokumentu jest pozniejsza niz zapostowania!"
    zlyallstar = "Allstar - zablokowac!"
    zlyairplus = "Airplus - zablokowac!"
    zlygoodyear = "GoodYear - zablokowac kredyt note!"
    zlacurrency = "Check Currency in vendor"

    dict_comments = {
        iter: {iter_column: ""}
    }
    # DODANIE ZAKLADKI ANALYSIS DO NOWEGO SLOWNIKA "DICT COMMENTS" ABY MIEC
    # W SLOWNIKU TEZ KOMENTARZE Z ERRORAMI



    for row in range(2, analysis_tab.max_row + 1):
        for col in "ABCDEFGHIJKLMNOPQ":
            cell = "{}{}".format(col, row)
            wartosc = analysis_tab[cell].value

            if iter_column == 1:
                dict_comments[iter] = {}

            dict_comments[iter][iter_column] = wartosc

            if iter_column < 17:
                iter_column = iter_column + 1
            else:
                iter_column = 1
                iter = iter + 1

    # WYPISANIE DANYCH ZE SLOWNIKA DO NOWEJ ZAKLADKI "DATA READY TO COPY"
    print("Sprawdzanie: czy dokumenty wystepuja w pliku 'dokumenty.txt'")

    if plikjestpusty == 0:
        f = open("dokumenty.txt", 'a+')
    else:
        f = open("dokumenty.txt", 'w+')

    for wolniej in dict_comments:
        if str(dict_comments[wolniej][2]) not in str(pliktxt): # JESLI NR DOKUMENTU ZNAJDUJE SIE W PLIKU TXT - OLEJ
            if dict_comments[wolniej][15] != None:
                if zladata in dict_comments[wolniej][15]:
                    data_tab.cell(column=3, row=nr_wiersza,
                                  value="not OK").font = style_data_tab  # WYPISANIE SPRAWDZONEJ DATY - ZLEJ

                    blad_data = 1  # JESLI WYSTAPIL BLAD CHOCIAZ RAZ - DAJ NOWA LINIE WIERSZA

                if zlyallstar in dict_comments[wolniej][15]:
                    data_tab.cell(column=13, row=nr_wiersza, value="not OK").font = style_data_tab
                    blad_allstar = 1

                if zlyairplus in dict_comments[wolniej][15]:
                    data_tab.cell(column=13, row=nr_wiersza, value="not OK").font = style_data_tab
                    blad_airplus = 1

            if dict_comments[wolniej][17] != None:
                if zlacurrency in dict_comments[wolniej][17]:
                    data_tab.cell(column=6, row=nr_wiersza, value="not OK").font = style_data_tab
                    blad_currency = 1

            suma_bledow = blad_currency + blad_allstar + blad_airplus + blad_data

            if suma_bledow > 0:  # JESLI CHOCIAZ RAZ WYSTAPIL BLAD, NASTEPNY WYPISZ W NOWEJ LINIJCE
                f.write(str(dict_comments[wolniej][2]) + "\n") # ZAPISANIE NR DOKUMENTU DO PLIKU
                data_tab.cell(column=1, row=nr_wiersza,
                              value=dict_comments[wolniej][2]).font = numerydok_data_tab  # WYPISANIE NR DOKUMENTU
                data_tab['A' + str(nr_wiersza)].fill = numerydok_data_tab_kolor # ZMIEN STYL NA NIEBIESKI DLA PIERWSZEJ KOLUMNY
                data_tab.cell(column=2, row=nr_wiersza, value="N/A").font = style_data_tab  # INV NUMBER

                if blad_data != 1:
                    data_tab.cell(column=3, row=nr_wiersza, value="OK").font = style_data_tab

                data_tab.cell(column=4, row=nr_wiersza, value="N/A").font = style_data_tab

                if blad_currency != 1:
                    data_tab.cell(column=6, row=nr_wiersza, value="OK").font = style_data_tab

                data_tab.cell(column=5, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=7, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=8, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=9, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=11, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=12, row=nr_wiersza, value="N/A").font = style_data_tab

                if (blad_airplus != 1) and (blad_allstar != 1):
                    data_tab.cell(column=13, row=nr_wiersza, value="OK").font = style_data_tab

                data_tab.cell(column=14, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=15, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=16, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=17, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=18, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=19, row=nr_wiersza, value="N/A").font = style_data_tab
                data_tab.cell(column=20, row=nr_wiersza, value="N/A").font = style_data_tab

                nr_wiersza = nr_wiersza + 1
                blad_airplus = 0  # ZRESETUJ LICZNIK BLEDOW
                blad_allstar = 0
                blad_currency = 0
                blad_currency = 0
                blad_data = 0


    # DODANIE NAGLOWKA DO ZAKLADKI "DATA READY TO COPY" + STYLE

    data_tab.cell(column=1, row=1, value="Document number").font = numerydok_data_tab
    data_tab.cell(column=2, row=1, value="Invoice number").font = naglowek_data_tab
    data_tab.cell(column=3, row=1, value="Invoice date").font = naglowek_data_tab
    data_tab.cell(column=4, row=1, value="Invoice sum").font = naglowek_data_tab
    data_tab.cell(column=5, row=1, value="Tax amount").font = naglowek_data_tab
    data_tab.cell(column=6, row=1, value="Currency").font = naglowek_data_tab
    data_tab.cell(column=7, row=1, value="Supplier data").font = naglowek_data_tab
    data_tab.cell(column=8, row=1, value="IBAN").font = naglowek_data_tab
    data_tab.cell(column=9, row=1, value="VW data").font = naglowek_data_tab
    data_tab.cell(column=11, row=1, value="Posting date").font = naglowek_data_tab
    data_tab.cell(column=12, row=1, value="Posting period").font = naglowek_data_tab
    data_tab.cell(column=13, row=1, value="Document type").font = naglowek_data_tab
    data_tab.cell(column=14, row=1, value="Payment block").font = naglowek_data_tab
    data_tab.cell(column=15, row=1, value="PO posting - price").font = naglowek_data_tab
    data_tab.cell(column=16, row=1, value="PO posting - quantity").font = naglowek_data_tab
    data_tab.cell(column=17, row=1, value="non PO posting - GL account").font = naglowek_data_tab
    data_tab.cell(column=18, row=1, value="non PO posting - cost center").font = naglowek_data_tab
    data_tab.cell(column=19, row=1, value="Vendor account").font = naglowek_data_tab
    data_tab.cell(column=20, row=1, value="Tax code").font = naglowek_data_tab

    data_tab['A1'].fill = numerydok_data_tab_kolor
    data_tab['B1'].fill = naglowek_data_tab_kolor
    data_tab['C1'].fill = naglowek_data_tab_kolor
    data_tab['D1'].fill = naglowek_data_tab_kolor
    data_tab['E1'].fill = naglowek_data_tab_kolor
    data_tab['F1'].fill = naglowek_data_tab_kolor
    data_tab['G1'].fill = naglowek_data_tab_kolor
    data_tab['H1'].fill = naglowek_data_tab_kolor
    data_tab['I1'].fill = naglowek_data_tab_kolor
    data_tab['K1'].fill = naglowek_data_tab_kolor
    data_tab['L1'].fill = naglowek_data_tab_kolor
    data_tab['M1'].fill = naglowek_data_tab_kolor
    data_tab['N1'].fill = naglowek_data_tab_kolor
    data_tab['O1'].fill = naglowek_data_tab_kolor
    data_tab['P1'].fill = naglowek_data_tab_kolor
    data_tab['Q1'].fill = naglowek_data_tab_kolor
    data_tab['R1'].fill = naglowek_data_tab_kolor
    data_tab['S1'].fill = naglowek_data_tab_kolor
    data_tab['T1'].fill = naglowek_data_tab_kolor



    data_tab.column_dimensions["A"].width = 23
    data_tab.row_dimensions[1].height = 123

    data_tab.column_dimensions["B"].width = 13
    data_tab.column_dimensions["C"].width = 13
    data_tab.column_dimensions["D"].width = 13
    data_tab.column_dimensions["E"].width = 13
    data_tab.column_dimensions["F"].width = 13
    data_tab.column_dimensions["G"].width = 13
    data_tab.column_dimensions["H"].width = 13
    data_tab.column_dimensions["I"].width = 13
    data_tab.column_dimensions["J"].width = 1
    data_tab.column_dimensions["K"].width = 17
    data_tab.column_dimensions["L"].width = 17
    data_tab.column_dimensions["M"].width = 17
    data_tab.column_dimensions["N"].width = 17
    data_tab.column_dimensions["O"].width = 17
    data_tab.column_dimensions["P"].width = 17
    data_tab.column_dimensions["Q"].width = 17
    data_tab.column_dimensions["R"].width = 17
    data_tab.column_dimensions["S"].width = 17
    data_tab.column_dimensions["T"].width = 17

    # ZMIANA STYLI DLA NAGLOWKA - ZAKLADKA ANALIZA
    style_naglowek = Font(bold=True, italic=False, name='Arial', size=9, color='FFFFFF')
    kolor_naglowek = PatternFill(start_color='003466', end_color='003466', fill_type='solid')
    kolor_quality = PatternFill(start_color='00182e', end_color='00182e', fill_type='solid')

    result_tab['A1'].fill = kolor_naglowek
    result_tab['B1'].fill = kolor_naglowek
    result_tab['C1'].fill = kolor_naglowek
    result_tab['D1'].fill = kolor_naglowek
    result_tab['E1'].fill = kolor_naglowek
    result_tab['F1'].fill = kolor_naglowek
    result_tab['G1'].fill = kolor_naglowek
    result_tab['H1'].fill = kolor_naglowek
    result_tab['I1'].fill = kolor_naglowek
    result_tab['J1'].fill = kolor_naglowek
    result_tab['K1'].fill = kolor_naglowek
    result_tab['L1'].fill = kolor_naglowek
    result_tab['M1'].fill = kolor_naglowek
    result_tab['N1'].fill = kolor_naglowek
    result_tab['O1'].fill = kolor_quality
    result_tab['P1'].fill = kolor_quality
    result_tab['Q1'].fill = kolor_quality


    result_tab['A1'].alignment = Alignment(wrap_text=True)
    result_tab['B1'].alignment = Alignment(wrap_text=True)
    result_tab['C1'].alignment = Alignment(wrap_text=True)
    result_tab['D1'].alignment = Alignment(wrap_text=True)
    result_tab['E1'].alignment = Alignment(wrap_text=True)
    result_tab['F1'].alignment = Alignment(wrap_text=True)
    result_tab['G1'].alignment = Alignment(wrap_text=True)
    result_tab['H1'].alignment = Alignment(wrap_text=True)
    result_tab['I1'].alignment = Alignment(wrap_text=True)
    result_tab['J1'].alignment = Alignment(wrap_text=True)
    result_tab['K1'].alignment = Alignment(wrap_text=True)
    result_tab['L1'].alignment = Alignment(wrap_text=True)
    result_tab['M1'].alignment = Alignment(wrap_text=True)
    result_tab['N1'].alignment = Alignment(wrap_text=True)
    result_tab['O1'].alignment = Alignment(wrap_text=True)
    result_tab['P1'].alignment = Alignment(wrap_text=True)
    result_tab['Q1'].alignment = Alignment(wrap_text=True)

    result_tab['A1'].font = style_naglowek
    result_tab['B1'].font = style_naglowek
    result_tab['C1'].font = style_naglowek
    result_tab['D1'].font = style_naglowek
    result_tab['E1'].font = style_naglowek
    result_tab['F1'].font = style_naglowek
    result_tab['G1'].font = style_naglowek
    result_tab['H1'].font = style_naglowek
    result_tab['I1'].font = style_naglowek
    result_tab['J1'].font = style_naglowek
    result_tab['K1'].font = style_naglowek
    result_tab['L1'].font = style_naglowek
    result_tab['M1'].font = style_naglowek
    result_tab['N1'].font = style_naglowek
    result_tab['O1'].font = style_naglowek
    result_tab['P1'].font = style_naglowek
    result_tab['Q1'].font = style_naglowek

    result_tab.column_dimensions["A"].width = 13
    result_tab.column_dimensions["B"].width = 9
    result_tab.column_dimensions["C"].width = 4
    result_tab.column_dimensions["D"].width = 5
    result_tab.column_dimensions["E"].width = 8
    result_tab.column_dimensions["F"].width = 9
    result_tab.column_dimensions["G"].width = 8
    result_tab.column_dimensions["H"].width = 7
    result_tab.column_dimensions["I"].width = 9
    result_tab.column_dimensions["J"].width = 9
    result_tab.column_dimensions["K"].width = 8
    result_tab.column_dimensions["L"].width = 9
    result_tab.column_dimensions["M"].width = 8
    result_tab.column_dimensions["N"].width = 7
    result_tab.column_dimensions["O"].width = 39
    result_tab.column_dimensions["P"].width = 24
    result_tab.column_dimensions["Q"].width = 28

    result_tab.row_dimensions[1].height = 38

    print("\nZapisywanie pliku...")

    #  print("\n")

    #### ZAPISZ PLIK ####
    book.save(plik_wsadowy+".xlsx")

    print("Zapisano plik "+str(plik_wsadowy)+".xlsx")

    # STOP CZAS - PART2
    end_time = time.time()

    # Obliczenie czasu wykonywania kodu i format bardziej przyjazny - minuty/godziny
    temp = end_time - start_time
    hours = temp // 3600
    temp = temp - 3600 * hours
    minutes = temp // 60
    seconds = temp - 60 * minutes
    print('Czas generowania raportu: %d:%d:%d' % (hours, minutes, seconds))

    koniec = input("Zakonczyc program? T/N")
    konieclow = koniec.lower()

    if konieclow == "t":
        close = "yes"
    else:
        close = "no"



